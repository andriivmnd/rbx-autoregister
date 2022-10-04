import os
import time
import openpyxl
import requests
import createdb as db

from selenium import webdriver
from selenium.webdriver.chrome.service import Service as ChromeService
from selenium.webdriver.support.ui import Select
from selenium.webdriver.common.by import By
from selenium.common import exceptions
from webdriver_manager.chrome import ChromeDriverManager

import config

options = webdriver.ChromeOptions()
options.add_argument('--disable-notifications')
options.add_argument("--window-size=300,850")

def is_cookie_valid(cookie):
    """Проверка через API роблокса работоспособность куки"""
    c = requests.Session()
    c.cookies[".ROBLOSECURITY"] = cookie

    try:
        response = c.post(
            url="https://auth.roblox.com/v1/authentication-ticket/",
            headers={
                "User-Agent": "RobloxStudio/WinInet",
                "Referer": "https://www.roblox.com/develop",
                "RBX-For-Gameauth": "true",
                "X-CSRF-TOKEN": getXsrfToken(cookie)[0]
            },
            allow_redirects=False,
            verify=False,
        )

        authcode = response.headers.get("rbx-authentication-ticket", None)
        if authcode == None:
            print("Invalid cookie!")
            return False
    except requests.exceptions.ConnectionError as e:
        print(f"ConnectionError - {e}")
        return False
    except Exception as e:
        print(f"Unknown exception:\nClass - {e.__class__}\nData - {e}")
        return False
    return True

XsrfTokenSaves = {}
def getXsrfToken(cookie, new=False):
    if not new and cookie in XsrfTokenSaves:
        print(f"old: {XsrfTokenSaves[cookie]}")
        return XsrfTokenSaves[cookie], True
    try:
        # В случае если пользователь не авторизован, то токен всё равно не нулевой.
        _request = requests.post("https://auth.roblox.com/v2/logout", cookies={".ROBLOSECURITY": cookie})
        token = _request.headers["x-csrf-token"]
        XsrfTokenSaves[cookie] = token
        print(f"new: {token}")
        return token, True
    except Exception as D:
        print("Error; getXsrfToken: " + str(D))
        return "", False

def buyVipServer(username, cookie):
    """Покупка вип сервера в adopt me!"""

    uri = "https://games.roblox.com/v1/games/vip-servers/383310974"

    while True:
        try:
            headers = {"X-CSRF-TOKEN": getXsrfToken(cookie)[0]}
            cookies = {".ROBLOSECURITY": cookie}
            data = {"expectedPrice": 0, "name": username}

            _request = requests.post(uri, headers=headers, cookies=cookies, data=data).json()

            uri_2 = f"https://games.roblox.com/v1/vip-servers/{_request['vipServerId']}"
            data_2 = {"newJoinCode": True}
            _request_2 = requests.patch(uri_2, headers=headers, cookies=cookies, data=data_2).json()

            return [_request["accessCode"], "https://www.roblox.com/games/920587237?privateServerLinkCode=" + _request_2["joinCode"]]
        except Exception as D:
            print("Exception!", D, cookie)

            cmd = input("[-] If you want try again: insert 0; to skip, insert 1:")
            while cmd not in ["1", "2"]:
                cmd = input("[-] If you want try again: insert 0; to skip, insert 1:")

            if cmd == "0":
                continue
            elif cmd == "1":
                break

def send_set_mail(username, cookie, mail, password):
    while True:
        try:
            headers = {"X-CSRF-TOKEN": getXsrfToken(cookie)[0]}
            cookies = {".ROBLOSECURITY": cookie}

            data = {
                'emailAddress': mail,
                'password': password,
            }

            _request = requests.post('https://accountsettings.roblox.com/v1/email', cookies=cookies, headers=headers, data=data).json()
            if "errors" in _request:
                if _request['errors'][0]['code'] == 4 or _request['errors'][0]['message'] == 'Challenge is required to authorize the request': #Если этот мейл уже стоит
                    return True
                elif _request['errors'][0]['message'] == "Token Validation Failed":
                    getXsrfToken(cookie, True)[0]
                    return False
                print(f"[MAIL] Roblox blocked our request: {_request['errors'][0]}")
                return False
            return True
        except Exception as D:
            print("Exception!", D, cookie)

            cmd = input("[-] If you want try again: insert 0; to skip, insert 1:")
            while cmd not in ["1", "2"]:
                cmd = input("[-] If you want try again: insert 0; to skip, insert 1:")

            if cmd == "0":
                continue
            elif cmd == "1":
                break

def set_2fa(id_profile, cookie):
    if not is_cookie_valid(cookie):
        return False

    headers = {"X-CSRF-TOKEN": getXsrfToken(cookie)[0]}
    cookies = {".ROBLOSECURITY": cookie}
    json_data = {}

    _request = requests.post(f'https://twostepverification.roblox.com/v1/users/{id_profile}/configuration/email/enable', cookies=cookies, headers=headers, json=json_data)
    if "errors" in _request.text:
        print(_request.text)
        return False
    return True

"""def set_pin(cookie, pin, password):
    if not is_cookie_valid(cookie):
        return False

    headers = {"X-CSRF-TOKEN": getXsrfToken(cookie)[0]}
    cookies = {".ROBLOSECURITY": cookie}

    response = requests.post('https://apis.roblox.com/reauthentication-service/v1/token/generate', cookies=cookies, headers=headers, json={
        'password': password,
    })

    token = response.status_code == 200 and response.json()["token"] or ""

    data = {
        'pin': pin,
        'reauthenticationToken': token,
    }

    _request = requests.post('https://auth.roblox.com/v1/account/pin', cookies=cookies, headers=headers, data=data)
    if "errors" in _request.json():
        print(_request.json(), data, response.json())
        return False
    return True"""

def register(sheet_login, sheet_password):
    #Регистрация
    service = ChromeService(executable_path=ChromeDriverManager().install())
    driver = webdriver.Chrome(service=service, options=options)
    driver.get('https://www.roblox.com/')
    month = Select(driver.find_element(By.ID, "MonthDropdown"))
    month.select_by_value("Jan")
    day = Select(driver.find_element(By.ID, "DayDropdown"))
    day.select_by_value("10")
    year = Select(driver.find_element(By.ID, "YearDropdown"))
    year.select_by_value("2000")
    
    username = driver.find_element(By.ID, "signup-username")
    username.send_keys(sheet_login)

    password = driver.find_element(By.ID, "signup-password")
    password.send_keys(sheet_password)

    gender = driver.find_element(By.ID, "MaleButton")
    gender.click()

    time.sleep(1)

    #доступен ли ник
    while driver.find_element(By.ID, "signup-usernameInputValidation").text != "": #проверка на допустимость ника в роблоксе
        sheet_login = db.gennick()

        username = driver.find_element(By.ID, "signup-username")
        username.clear()
        username.send_keys(sheet_login)
        time.sleep(1)
        
    signup = driver.find_element(By.ID, "signup-button")
    signup.click()
    
    #Проверяем есть ли капча, если есть то вводим вручную
    time.sleep(3)
    while True:
        time.sleep(1)
        try:
            driver.find_element(By.ID, "challenge-captcha-element")
        except exceptions.NoSuchElementException:
            break

    while True:
        if driver.current_url == "https://www.roblox.com/home?nu=true":  #Ждём пока регистрация пройдёт успешно
            break

    cookie = driver.get_cookie(".ROBLOSECURITY")["value"]

    infoTb = buyVipServer(sheet_login, cookie)

    return sheet_login, cookie, infoTb[1], infoTb[0]

def function_2fa(login, password, pin, cookie, step):
    service = ChromeService(executable_path=ChromeDriverManager().install())
    driver = webdriver.Chrome(service=service, options=options)

    driver.get('https://www.roblox.com/') #Заходим на страницу чтобы задать куки

    #Проверяем прошла ли авторизация успешно
    if (step == 1 or step == 4) and cookie and is_cookie_valid(cookie): #driver.current_url == "https://www.roblox.com/home":
        driver.add_cookie({"name":".ROBLOSECURITY","value":cookie,"path":"/","domain":"roblox.com"})
        driver.get('https://www.roblox.com/')
    else: #Авторизация если кука неправильная
        driver.get('https://www.roblox.com/login')
        
        _input = driver.find_element(By.ID, "login-username")
        _input.send_keys(login)

        _input = driver.find_element(By.ID, "login-password")
        _input.send_keys(password)

        time.sleep(1)

        signup = driver.find_element(By.ID, "login-button")
        signup.click()

        #Ждём капчу, если нет то идём дальше
        time.sleep(3)
        while True:
            time.sleep(1)
            try:
                driver.find_element(By.ID, "challenge-captcha-element")
                time.sleep(1)
                driver.find_element(By.ID, "challenge-captcha-element")
                print("challenge-captcha-element")
            except exceptions.NoSuchElementException:
                print("Капчу не нашел, или её уже ввели")
                break

        time.sleep(3)
        while True:
            time.sleep(1)
            try:
                print("two-step-verification-code-input")
                _input = driver.find_element(By.ID, "two-step-verification-code-input")
                code = getLastMail_info(login, password, "confirmLogin")

                _input.send_keys(code)

                button = driver.find_element(By.XPATH, "//button[@type='button'][@class='btn-cta-md modal-modern-footer-button']")
                button.click()
                break
            except:
                print("двухфакторки не нашел или она уже введена")
                break

        #Ждём пока авторизация пройдёт успешно
        while True:
            if driver.current_url == "https://www.roblox.com/home":
                break
    
    cookie = driver.get_cookie(".ROBLOSECURITY")["value"]

    #MAIL
    if step <= 2:
        print("MAIL")
        step = 2
        mail = config.TEST_MAIL and config.TEST_MAIL[0] or f"{login}@rbxspace.gg"
            
        while True:
            if send_set_mail(login, cookie, mail, password):
                driver.get('https://www.roblox.com/my/account#!/info')
                while True:
                    try:
                        #Подтвержение почты
                        driver.find_element(By.XPATH, "//span[@class='account-field-email-verify-msg text-error']")
                        status = getLastMail_info(login, password, "confirmMail")
                        break
                    except:
                        try:
                            span = driver.find_element(By.XPATH, "//span[@class='text-robux account-field-email-verified-text ng-binding']").text
                            if span == "Verified":
                                break
                        except:
                            pass
                break
            else:
                input("[ERROR] Press enter for continue\n")

    #2FA
    if step <= 3:
        print("2FA")
        step = 3
        #userid = driver.find_element(By.XPATH,"//meta[@name='user-data']").get_attribute("data-userid")
        #if not set_2fa(userid, cookie):
            #return False, step

        while True:
            driver.get('https://www.roblox.com/my/account#!/security')
            time.sleep(2)

            if driver.current_url != "https://www.roblox.com/my/account#!/security":
                return False, step

            try:
                btn = driver.find_element(By.ID, "2sv-toggle")
                if btn.get_attribute('class') != "btn-toggle receiver-destination-type-toggle on":
                    btn.click()
                    time.sleep(1)

                    try:
                        _input = driver.find_element(By.ID, "reauthentication-password-input")
                        _input.send_keys(password)
                        button = driver.find_element(By.XPATH, "//button[@type='button'][@class='btn-cta-md modal-modern-footer-button']")
                        button.click()
                        time.sleep(1)
                    except:
                        pass

                    if btn == "btn-toggle receiver-destination-type-toggle on":
                        break
                else:
                    break
            except:
                pass

    #PIN
    if step <= 4:
        print("PIN")
        step = 4
        #if not set_pin(cookie, pin, password):
            #return False, step

        #Проверка пинкода
        while True:
            driver.get('https://www.roblox.com/my/account#!/parental-controls')
            time.sleep(2)

            if driver.current_url != "https://www.roblox.com/my/account#!/parental-controls":
                return False, step

            try:
                span = driver.find_element(By.ID, "accountPin-toggle").get_attribute('class')
                if span != "btn-toggle receiver-destination-type-toggle on":
                    button = driver.find_element(By.ID, "accountPin-toggle")
                    button.click()
                    time.sleep(1)

                    _input = driver.find_element(By.XPATH, "//input[@name='newPin']")
                    _input.send_keys(pin)
                    _input = driver.find_element(By.XPATH, "//input[@name='newPinConfirm']")
                    _input.send_keys(pin)

                    button = driver.find_element(By.XPATH, "//button[@type='submit'][@class='modal-button btn-secondary-md ng-binding']")
                    button.click()
                    time.sleep(1)

                    try:
                        print("two-step-verification-code-input")
                        _input = driver.find_element(By.ID, "two-step-verification-code-input")
                        code = getLastMail_info(login, password, "confirmLogin")

                        _input.send_keys(code)

                        button = driver.find_element(By.XPATH, "//button[@type='button'][@class='btn-cta-md modal-modern-footer-button']")
                        button.click()
                        time.sleep(1)
                    except:
                        print("двухфакторки не нашел или она уже введена")
                        pass

                    try:
                        _input = driver.find_element(By.ID, "reauthentication-password-input")
                        _input.send_keys(password)
                        button = driver.find_element(By.XPATH, "//button[@type='button'][@class='btn-cta-md modal-modern-footer-button']")
                        button.click()
                        time.sleep(1)
                    except:
                        pass

                    span = driver.find_element(By.ID, "accountPin-toggle").get_attribute('class')
                    if span == "btn-toggle receiver-destination-type-toggle on":
                        break
                else: 
                    break
            except:
                pass

    if step <= 5:
        step = 5
        if not is_cookie_valid(cookie):
            return False, step

    cookie = driver.get_cookie(".ROBLOSECURITY")["value"]

    return True, cookie


session_saves_mail = {}
def getLastMail_info(login, password, wait):
    service = ChromeService(executable_path=ChromeDriverManager().install())
    driver = webdriver.Chrome(service=service, options=options)

    #Авторизация в почте
    driver.get(f'http://{config.URL_MAIL}/roundcubemail')
    mail_login = config.TEST_MAIL and config.TEST_MAIL[0] or f"{login}@rbxspace.gg"
    if mail_login in session_saves_mail:
        driver.add_cookie({"name":"language","value":"ru","path":"/","domain":config.URL_MAIL})
        driver.add_cookie({"name":"roundcube_sessauth","value":session_saves_mail[mail_login][0],"path":"/","domain":config.URL_MAIL})
        driver.add_cookie({"name":"roundcube_sessid","value":session_saves_mail[mail_login][1],"path":"/","domain":config.URL_MAIL})
        driver.get(f'http://{config.URL_MAIL}/roundcubemail')

        while True:
            if driver.current_url == f"http://{config.URL_MAIL}/roundcubemail/?_task=mail&_mbox=INBOX":
                break
    else:
        time.sleep(2)

        _input = driver.find_element(By.ID, "rcmloginuser")
        _input.send_keys(mail_login)

        _input = driver.find_element(By.ID, "rcmloginpwd")
        _input.send_keys(config.TEST_MAIL and config.TEST_MAIL[1] or password)

        button = driver.find_element(By.ID, "rcmloginsubmit")
        button.click()

        #Ждём пока авторизация пройдёт успешно
        while True:
            if driver.current_url == f"http://{config.URL_MAIL}/roundcubemail/?_task=mail&_mbox=INBOX":
                break

        session_saves_mail[mail_login] = [driver.get_cookie("roundcube_sessauth")["value"],driver.get_cookie("roundcube_sessid")["value"]]

    #Проверяем последнее сообщение на почте
    while True:
        try:
            #Получение кол-ва писем на ящике
            cookie_sessid = driver.get_cookie("roundcube_sessid")["value"]
            cookie_sessauth = driver.get_cookie("roundcube_sessauth")["value"]
            response = requests.get(f'http://{config.URL_MAIL}/roundcubemail/', params={'_task': 'mail','_action': 'list','_refresh': '1','_layout': 'widescreen','_mbox': 'INBOX','_remote': '1',}, cookies={
                'roundcube_sessid': cookie_sessid,
                'roundcube_sessauth': cookie_sessauth,
            }, verify=False)

            #Переходим на страницу последнего письма
            driver.get(f'http://{config.URL_MAIL}/roundcubemail/?_uid={response.json()["env"]["messagecount"]}&_framed=1&_action=preview')
            time.sleep(1)

            if wait == "confirmMail":
                #Проверяем логин из письма
                textd = driver.find_element(By.XPATH, "//h2[@class='subject']").text
                textd_login = textd[textd.find(':')+2 : len(textd)]
                if textd_login == login:
                    #Кликаем по ссылке с подтверждением
                    href = driver.find_element(By.XPATH, "//a[@class='v1email-button']").get_attribute("href")
                    driver.get(href)
                    while True:
                        time.sleep(1)
                        try:
                            driver.find_element(By.ID, "verify-email-container")
                            time.sleep(2)
                            break
                        except:
                            pass
                    return True
            elif wait == "confirmLogin":
                textd = driver.find_element(By.XPATH, "//h4[@class='v1email-header']")
                textd = textd.text
                textd = [s for s in textd.split() if s.isdigit()][0]
                print(textd)
                return textd
        except:
            time.sleep(3) #Ждём пока будет нужное письмо

def function_register(name, row_start=1):
    excel_book = openpyxl.load_workbook(name)
    sheet_obj = excel_book.active
    for i in range(row_start, sheet_obj.max_row+1):
        print(f"reg. Started the row: {i}")
        sheet_login = sheet_obj.cell(row = i, column = 1)
        sheet_password = sheet_obj.cell(row = i, column = 2)
        sheet_pin = sheet_obj.cell(row = i, column = 4)
        sheet_url = sheet_obj.cell(row = i, column = 5)
        sheet_accesscode = sheet_obj.cell(row = i, column = 6)
        sheet_cookie = sheet_obj.cell(row = i, column = 7)

        nickname, cookie, url, accesscode = register(sheet_login.value, sheet_password.value)
        sheet_login.value = nickname
        sheet_cookie.value = cookie
        sheet_url.value = url
        sheet_accesscode.value = accesscode
        print(f"reg. Finished the row: {i}")
        excel_book.save(name)

def function_login(name, row_start=1):
    excel_book = openpyxl.load_workbook(name)
    sheet_obj = excel_book.active
    for i in range(row_start, sheet_obj.max_row+1):
        print(f"2fa. Started the row: {i}")
        sheet_login = sheet_obj.cell(row = i, column = 1)
        sheet_password = sheet_obj.cell(row = i, column = 2)
        sheet_pin = sheet_obj.cell(row = i, column = 4)
        sheet_url = sheet_obj.cell(row = i, column = 5)
        sheet_accesscode = sheet_obj.cell(row = i, column = 6)
        sheet_cookie = sheet_obj.cell(row = i, column = 7)
        sheet_ifoldstep = sheet_obj.cell(row = i, column = 8)

        if sheet_login and sheet_ifoldstep.value != "true":
            step = 1
            while True:
                status, cookie = function_2fa(sheet_login.value, sheet_password.value, sheet_pin.value, sheet_cookie.value, step)
                if status == True:
                    sheet_cookie.value = cookie
                    sheet_ifoldstep.value = "true"
                    print(f"2fa. Finished the row: {i}")
                    break
                else:
                    print(f"STEP: {cookie}")
                    step = cookie


        excel_book.save(name)

def checkPin(name, row_start=1):
    excel_book = openpyxl.load_workbook(name)
    sheet_obj = excel_book.active
    for i in range(row_start, sheet_obj.max_row+1):
        print(f"pin. Started the row: {i}")
        sheet_login = sheet_obj.cell(row = i, column = 1)
        sheet_password = sheet_obj.cell(row = i, column = 2)
        sheet_pin = sheet_obj.cell(row = i, column = 4)
        sheet_url = sheet_obj.cell(row = i, column = 5)
        sheet_accesscode = sheet_obj.cell(row = i, column = 6)
        sheet_cookie = sheet_obj.cell(row = i, column = 7)

        if sheet_login:
            step = 4
            while True:
                status, cookie = function_2fa(sheet_login.value, sheet_password.value, sheet_pin.value, sheet_cookie.value, step)
                if status == True:
                    sheet_cookie.value = cookie
                    print(f"pin. Finished the row: {i}")
                    break
                else:
                    print(f"STEP: {cookie}")
                    step = cookie


        excel_book.save(name)



def main():
    for dir in ["output"]:
        if not os.path.exists(dir):
            os.makedirs(dir)

    print("\nВыбор функционала:\n1. Создание аккаунтов(с генерацией)\n2. Создание аккаунтов(без генерации)\n3. Включение 2fa\n4. Проверка и установка пинкода\n")
    select = input("Select function: ")
    
    if select == "1":
        name = "output/" + input("File name: ") + ".xlsx"
        db.create(name)
        if os.path.exists(name):
            function_register(name)

    elif select == "2":
        name = "output/" + input("Select file name: ") + ".xlsx"
        row_start = int(input("Select row start: "))
        if os.path.exists(name):
            function_register(name, row_start)
        else:
            print(f"No exists file {name}")

    elif select == "3":
        name = "output/" + input("Select file name: ") + ".xlsx"
        row_start = int(input("Select row start: "))
        #row_count = int(input("Select row count: "))
        if os.path.exists(name):
            function_login(name, row_start)
        else:
            print(f"No exists file {name}")

    elif select == "4":
        name = "output/" + input("Select file name: ") + ".xlsx"
        row_start = int(input("Select row start: "))
        if os.path.exists(name):
            checkPin(name, row_start)
        else:
            print(f"No exists file {name}")

    main()

if __name__ == '__main__':
    main()